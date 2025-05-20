# import packages and dependencies

import requests
import pandas as pd
import threading
import json
import openpyxl
from openpyxl import load_workbook
import xlsxwriter

import dependency_function_pdc
from dependency_function_pdc import query_pdc # function to retrieve data
import dependency_query # import queries list and subfunctions
from dependency_query import *
import dependency_headers
from dependency_headers import *
import dependency_for_loops
from dependency_for_loops import *

# only for testing, introduce pdc_study_id input
try:
    pdc_study_id_input = str(input("Introduce a PDC study id, ie: PDC000443: "))
except ValueError:
    print("Please enter a valid id. PDC study ids are available at https://pdc.cancer.gov/")

# Define Variables

# Variables
variables = {
    "pdc_study_identifier": pdc_study_id_input,
    "offset": 0, 
    "limit": 20700
}
workbook_data = {}

def block_case_sample():
    variables_case = variables.copy()
    # Case-Matrix
    speciment_data = query_pdc(query= query_biospecimen, variables=variables_case)
    matrix = json.loads(speciment_data.content)['data']["biospecimenPerStudy"]
    biospecimen_df = pd.DataFrame(matrix)

    # Case
    case_data = query_pdc(query= query_case, variables=variables_case)
    matrix = json.loads(case_data.content)['data']['case']
    case_data_df = pd.DataFrame(matrix)
    case = pd.merge(left=biospecimen_df, right=case_data_df, on="case_id")
    columns_to_keep = [col for col in case.columns if not col.endswith('_y')]
    case = case[columns_to_keep]
    case = case.rename(columns={col: col.rstrip('_x') for col in case.columns})
    to_remove = list(set(case.columns).difference(case_header))
    case.drop(columns=to_remove, inplace=True)
    case = case.reindex(columns=case_header)

    # Sample
    #matrix = json.loads(case_data.content)['data']['case']
    #case_data_df = pd.DataFrame(matrix)
    sample_df = for_sample(matrix = matrix)
    to_remove = list(set(sample_df.columns).difference(sample_header))
    samples = sample_df.drop(columns=to_remove)
    samples = samples.reindex(columns=sample_header)
    
    # Expanded version of case-matrix 
    tmp = pd.merge(
        left=biospecimen_df,
        right=sample_df,
        left_on="sample_id",
        right_on="sample_id",
        indicator=True
    )
    tmp = tmp.rename(columns={col: col.rstrip('_x') for col in tmp.columns})
    new_columns = [
        "aliquot_id", "aliquot_submitter_id", "sample_id", "sample_submitter_id",
        "case_id","case_submitter_id", "sample_type",
        "disease_type", "primary_site", "tissue_type"
    ]

    to_remove = list(set(tmp.columns).difference(new_columns))
    tmp = tmp.drop(columns=to_remove)

    case_matrix = tmp.reindex(columns=new_columns)

    # Aliquots
    aliquots_data = query_pdc(query= query_aliquots, variables=variables_case)
    matrix = json.loads(aliquots_data.content)["data"]["paginatedCasesSamplesAliquots"]["casesSamplesAliquots"]
    aliquots_df = pd.DataFrame(matrix)
    aliquots_df['gdc_sample_id'] = aliquots_df['samples'].apply(lambda diag_list: diag_list[0]['gdc_sample_id'] if diag_list else None)
    tmp = for_aliquots(matrix= matrix)
    aliquots_df = pd.merge(left=biospecimen_df, right=tmp, on = "aliquot_submitter_id",suffixes= ("", "_"))
    to_remove = list(set(aliquots_df.columns).difference(aliquots_header))
    aliquots_df = aliquots_df.drop(columns=to_remove)
    aliquots = aliquots_df.reindex(columns=aliquots_header)

    workbook_data["case_matrix"]= case_matrix 
    workbook_data["case"]= case 
    workbook_data["samples"]= samples
    workbook_data["aliquots"]= aliquots

    print("end of block case-matrix")

def block_clinical():
    variables_clinical = variables.copy()
    # Project-Program 
    study_data = query_pdc(query= query_study_info, variables=variables_clinical)
    matrix = json.loads(study_data.content)['data']['study']
    study_df = pd.DataFrame(matrix)
    program_project = study_df[program_project_header].transpose()
    program_project.columns = ['name']
    program_project[" "] = program_project.index
    program_project = program_project.reindex(columns=[' ', 'name'])

    # Demographic

    variables_clinical['study_id'] = study_df['study_id'][0]
    demographics_data = query_pdc(query= query_demographcis, variables=variables_clinical)
    matrix = json.loads(demographics_data.content)['data']["paginatedCaseDemographicsPerStudy"]["caseDemographicsPerStudy"]
    demographics_data = pd.DataFrame(matrix[1:], columns=matrix[0])
    demographics_data['demographic_id'] = demographics_data['demographics'].apply(lambda diag_list: diag_list[0]['demographic_id'] if diag_list else None)
    demographics_df = for_demographics(matrix = matrix)
    demographic = pd.merge(left=demographics_data, right=demographics_df, on="demographic_id")
    to_remove = list(set(demographic.columns).difference(demographics_header))
    demographic.drop(columns=to_remove, inplace=True)
    demographic = demographic.reindex(columns=demographics_header)


    # Diganosis
    diagnose_data = query_pdc(query= query_diagnose, variables=variables_clinical)
    matrix = json.loads(diagnose_data.content)['data']["paginatedCaseDiagnosesPerStudy"]["caseDiagnosesPerStudy"]
    diagnose_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
    diagnose_data_df['diagnosis_id'] = diagnose_data_df['diagnoses'].apply(lambda diag_list: diag_list[0]['diagnosis_id'] if diag_list else None)
    diagnose_df = for_diagnosis(matrix = matrix)
    diagnosis = pd.merge(left=diagnose_data_df, right=diagnose_df, on="diagnosis_id")
    to_remove = list(set(diagnosis.columns).difference(diagnose_header))
    diagnosis.drop(columns=to_remove, inplace=True)
    diagnosis = diagnosis.reindex(columns=diagnose_header)


    # Exposure

    exposure_data = query_pdc(query= query_exposure, variables= variables_clinical)
    matrix = json.loads(exposure_data.content)['data']["paginatedCaseExposuresPerStudy"]["caseExposuresPerStudy"]
    exposure_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
    exposure_data_df['exposure_id'] = exposure_data_df['exposures'].apply(lambda diag_list: diag_list[0]['exposure_id'] if diag_list else None)
    exposure_df = for_exposure(matrix = matrix)
    exposure = pd.merge(left=exposure_data_df, right=exposure_df, on="exposure_id")
    to_remove = list(set(exposure.columns).difference(expose_header))
    exposure = exposure.drop(columns=to_remove)
    exposure = exposure.reindex(columns=expose_header)

    # Treatment

    treatments_data = query_pdc(query= query_treatments, variables=variables_clinical)
    matrix = json.loads(treatments_data.content)['data']["paginatedCaseTreatmentsPerStudy"]["caseTreatmentsPerStudy"]
    treatments_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
    treatments_data_df['treatment_id'] = treatments_data_df['treatments'].apply(lambda diag_list: diag_list[0]['treatment_id'] if diag_list else None)
    treatments_df = for_treatment(matrix = matrix)
    treatments = pd.merge(left=treatments_data_df, right=treatments_df, on="treatment_id")
    to_remove = list(set(treatments.columns).difference(treatment_header))
    treatments = treatments.drop(columns=to_remove)
    treatments = treatments.reindex(columns=treatment_header)


    # Follow up

    follow_up_data = query_pdc(query= query_follow_up, variables=variables_clinical)
    matrix = json.loads(follow_up_data.content)['data']["paginatedCaseFollowUpsPerStudy"]['caseFollowUpsPerStudy']
    follow_up_data_df = pd.DataFrame(matrix[1:], columns=matrix[0])
    follow_up_data_df['follow_up_id'] = follow_up_data_df['follow_ups'].apply(lambda diag_list: diag_list[0]['follow_up_id'] if diag_list else None)
    follow_up_df = for_follows_up(matrix = matrix)
    follow_ups = pd.merge(left=follow_up_data_df, right=follow_up_df, on="follow_up_id")
    to_remove = list(set(follow_ups.columns).difference(follow_ups_header))
    follow_ups = follow_ups.drop(columns=to_remove)
    follow_ups = follow_ups.reindex(columns=follow_ups_header)


    # Study - Data
    matrix = json.loads(study_data.content)['data']['study']
    study_df = pd.DataFrame(matrix)
    to_remove = list(set(study_df.columns).difference(study_header))
    study = study_df.drop(columns=to_remove)
    study = study.reindex(columns=study_header)


    # Protocol
    protocol_Data = query_pdc(query= query_protocol, variables=variables_clinical)
    matrix = json.loads(protocol_Data.content)['data']['protocolPerStudy']
    protocol_df = pd.DataFrame(matrix)
    to_remove = list(set(protocol_df.columns).difference(protocol_header))
    protocol = protocol_df.drop(columns=to_remove)
    protocol = protocol.reindex(columns=protocol_header)
    
    workbook_data["program_project"]= program_project 
    workbook_data["demographic"]= demographic 
    workbook_data["diagnosis"] = diagnosis
    workbook_data["exposure"] = exposure
    workbook_data["treatments"] = treatments
    workbook_data['follow_ups'] = follow_ups
    workbook_data["study"] = study
    workbook_data["study_df"]=study_df
    workbook_data["protocol"] = protocol
    print("end of block clinical")

def block_metadata():
    variables_meta = variables.copy()
    # Experimental - Metadata
    expMetadat_data_2 = query_pdc(query= query_expMetadata_2, variables=variables_meta)
    matrix = json.loads(expMetadat_data_2.content)['data']["studyExperimentalDesign"]
    expMetadat_data_2 = pd.DataFrame(matrix)
    to_remove = list(set(expMetadat_data_2.columns).difference(exp_metadata_header))
    Exp_Metadata = expMetadat_data_2.drop(columns=to_remove)
    Exp_Metadata = Exp_Metadata.reindex(columns=exp_metadata_header)

    #tmt_columns = [
    #    'tmt_126', 'tmt_127n', 'tmt_127c', 'tmt_128n', 'tmt_128c',
    #    'tmt_129n', 'tmt_129c', 'tmt_130n', 'tmt_130c', 'tmt_131', 'tmt_131c'
    #]

    # Function to extract aliquot_ids
    #def extract_aliquots(cell):
     #   if isinstance(cell, list):
      #      return [entry.get('aliquot_id') for entry in cell if isinstance(entry, dict)]
       # return []

    # Function to format aliquot_ids
    #def format_aliquot_ids(aliquot_list):
    #    return [f"aliquot_id: {aliquot}" for aliquot in aliquot_list]

    # Modify Exp_Metadata in-place
    #for col in tmt_columns:
    #    if col in Exp_Metadata.columns:
     #       Exp_Metadata[col] = Exp_Metadata[col].apply(extract_aliquots).map(format_aliquot_ids)

    
    # File Metadata

    file_metadata_data = query_pdc(query= query_file_metadata, variables=variables_meta)
    matrix = json.loads(file_metadata_data.content)['data']["filesPerStudy"]
    file_metadata_df = pd.DataFrame(matrix)
    to_remove = list(set(file_metadata_df.columns).difference(file_metadata_header))
    file_metada = file_metadata_df.drop(columns=to_remove)
    file_metada = file_metada.reindex(columns=file_metadata_header)


    file_id_list = file_metadata_df["file_id"].dropna().astype(str).tolist()


    all_metadata = []

    # Loop through each file_id
    for file_id in file_id_list:
        variables_2 = {
            "file_id": file_id,
            "offset": 0,
            "limit": 20700
        }
        try:
            result = query_pdc(query=query_file_metadata_2, variables=variables_2)
            matrix = json.loads(result.content)['data']["fileMetadata"]
            all_metadata.extend(matrix)
        except Exception as e:
            print(f"Failed for file_id {file_id}: {e}")

    all_metadata_df = pd.DataFrame(all_metadata)

    tmp = pd.merge(file_metadata_df, all_metadata_df, on="file_id")
    columns_to_keep = [col for col in tmp.columns if not col.endswith('_y')]
    tmp = tmp[columns_to_keep]
    tmp = tmp.rename(columns={col: col.rstrip('_x') for col in tmp.columns})
    tmp = tmp.rename(columns={'plex_or_dataset_name': "plex_or_folder_name", 
                        "study_run_metadata_id": "run_metadata_id",
                        "protocol_submitter_id": "protocol"})
    to_remove = list(set(tmp.columns).difference(file_metadata_header_2))
    file_metada = tmp.drop(columns=to_remove)
    file_metada = file_metada.reindex(columns=file_metadata_header_2)

    workbook_data["Exp_Metadata"] = Exp_Metadata
    workbook_data["file_metada"] = file_metada
    print("end of block metadata")

def block_workbook():
    print("preparing workbook")
    # object dictionary:
    study_information = {
        #'Readme': pd.DataFrame(readme),
        "Project-Program": workbook_data.get('program_project'),
        "Case_Matrix": workbook_data.get('case_matrix'),
        "Case": workbook_data.get('case'),
        "Demographic": workbook_data.get("demographic"),
        "Diagnosis": workbook_data.get("diagnosis"),
        "Exposure": workbook_data.get('exposure'),
        "Family History": pd.DataFrame(), #need to find the data
        "Treatment": workbook_data.get("treatments"),
        "Follow-up": workbook_data.get("follow_ups"),
        "Sample": workbook_data.get("samples"), #need to redo
        "Aliquots": workbook_data.get("aliquots"), #need to redo
        "Study": workbook_data.get("study"),
        "Protocol": workbook_data.get("protocol"),
        "Exp_Runmetadata": workbook_data.get("Exp_Metadata"),
        "File_Metadata": workbook_data.get("file_metada")
    }

    filename = f"Study_Data_{pdc_study_id_input}.xlsx"

    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        for sheet_name, df in study_information.items():
            if df is not None and not df.empty:  # Check if the DataFrame is not None and not empty
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Create an empty DataFrame and write it to the sheet
                empty_df = pd.DataFrame(columns= df.columns, index=range(10)).fillna('data not available')
                empty_df.to_excel(writer, sheet_name=sheet_name, index=False)


    # File path to the existing Excel file

    # Prepare your two DataFrames
    study_df = workbook_data.get('study_df')
    study_transposed = study_df.transpose()
    study_transposed = study_transposed[:-1]

    files_count_df = pd.DataFrame(study_df.loc[0, "filesCount"])

    # Define a unique new sheet name
    sheet_name = "Study Summary"

    # Open writer to append a sheet (no custom 'book' logic)
    with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        # Write the transposed study_df
        study_transposed.to_excel(writer, index=True, header=True, sheet_name=sheet_name, startrow=0)
        
        # Leave a gap and write the transposed filesCount
        start_row = study_transposed.shape[0] + 5
        files_count_df.to_excel(writer, index=False, header=True, sheet_name=sheet_name, startrow=start_row)

    wb = load_workbook(filename)
    sheet_order = wb.sheetnames
    sheet_order.insert(0, sheet_order.pop(sheet_order.index("Study Summary")))
    wb._sheets = [wb[sheet] for sheet in sheet_order]

    wb.save(filename)

    # almost complete, need to remove the last row from study_transpiosed and positioned to the beggining
    print('download completed')

def main():
    threads = []

    t1 = threading.Thread(target=block_case_sample)
    t2 = threading.Thread(target=block_clinical)
    t3 = threading.Thread(target=block_metadata)

    threads.extend([t1, t2, t3])

    for t in threads:
        t.start()
    for t in threads:
        t.join()
    block_workbook()

if __name__ == "__main__":
    main()
